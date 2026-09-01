"""Builds TolStack_Template.xlsx with one example tab exercising every input."""
import os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Example"
BOLD=Font(bold=True); WHITE=Font(bold=True,color="FFFFFF")
def fill(c): return PatternFill("solid", fgColor=c)
HDR=fill("305496"); SEC=fill("8EAADB")
GREEN=fill("C6EFCE"); RED=fill("FFC7CE"); BLUE=fill("BDD7EE"); GOLD=fill("FFE699"); PURPLE=fill("E4DFEC")
thin=Side(style="thin",color="B0B0B0"); border=Border(thin,thin,thin,thin)
center=Alignment(horizontal="center")

def row(r, a=None,b=None,vals=None,note=None,f=None,font=None):
    if a is not None: ws.cell(r,1,a)
    if b is not None: ws.cell(r,2,b)
    if vals is not None:
        for j,v in enumerate(vals): ws.cell(r,3+j,v)
    if note is not None: ws.cell(r,9,note)
    if f:
        for c in range(1,9): ws.cell(r,c).fill=f
    if font:
        for c in range(1,9): ws.cell(r,c).font=font

# Title
ws.cell(1,1,"TolStack — Example Stack").font=Font(bold=True,size=16)
ws.cell(2,1,"Fill the yellow/green/blue cells, highlight A4:I31, then click Solve.").font=Font(italic=True,color="666666")

hdr=["Vector Name","Vector Tag","X","Y","Z","Tx","Ty","Tz","Notes"]
def header_row(r, title):
    row(r, title, f=SEC, font=WHITE)
    r+=1
    for j,h in enumerate(hdr): ws.cell(r,1+j,h)
    row(r, f=HDR, font=WHITE)
    return r+1

# NOMINAL
r=4; r=header_row(r,"NOMINAL VECTOR PATH  (positions mm, angles rad)")
row(r,"R1","R",[10,0,0,0,0,0],"10 mm along X",GREEN); r+=1
row(r,"R2","R",[0,10,0,0,0,0],"10 mm along Y",GREEN); r+=1
row(r,"R3","R",[0,0,5,0,0,0],"5 mm along Z",GREEN); r+=1
# ERROR
r=header_row(r,"RANDOM ERROR TERMS  (± value at N_SIGMA)")
row(r,"R1e","Re",[0.02,0.02,0,0,0,0.001],"pos/ang tol at joint 1",RED); r+=1
row(r,"R2e","Re",[0.02,0.02,0,0,0,0],"",RED); r+=1
row(r,"R3e","Re",[0,0,0.03,0.0005,0.0005,0],"",RED); r+=1
# COMPENSATOR
r=header_row(r,"COMPENSATOR  (C=setpoint, Cv=lever vector, Ce=repeatability ±@Nσ)")
row(r,"C1","C",[0,0,0,0,0,0.0005],"nulls Tz toward setpoint",BLUE); r+=1
row(r,"C1v","Cv",[10,0,0,0,0,0],"corrector 10 mm away",BLUE); r+=1
row(r,"C1e","Ce",[0,0,0,0,0,0.0002],"corrector zero-mean error",BLUE); r+=1
# CONFIG
row(r,"MONTECARLO CONFIG",f=SEC,font=WHITE); r+=1
for name,tag,val,note in [("MC Samples","N_SAMPLES",10000,"# of trials (>=2)"),
                          ("Sigma","N_SIGMA",3,"tolerance is at this many sigma"),
                          ("Distribution","DISTRIBUTION","N","N=Normal, U=Uniform, T=Triangular"),
                          ("Seed","SEED",None,"blank = new random draw each run; a number = reproducible"),
                          ("Make Plots","PLOT",1,"1=embed plots in sheet, 0=none"),
                          ("Open Plot Windows","SHOW",0,"1=also open interactive (rotatable) windows"),
                          ("Stack Name","NAME","Example","label for plots")]:
    ws.cell(r,1,name); ws.cell(r,2,tag)
    if val is not None: ws.cell(r,3,val)
    ws.cell(r,9,note)
    for c in range(1,9): ws.cell(r,c).fill=GOLD
    r+=1
# RESULTS
r=header_row(r,"SIMULATION RESULTS  (written by Solve)")
for name,tag,note in [("mu + Nσ (spec)","RESULT","upper estimate mean+Nσ"),
                      ("Mean","MU","systematic bias"),
                      ("Sigma (1σ)","SIGMA","random spread"),
                      ("|mu| + Nσ (worst)","WORST_CASE","conservative worst case")]:
    ws.cell(r,1,name); ws.cell(r,2,tag); ws.cell(r,9,note)
    for c in range(1,9): ws.cell(r,c).fill=PURPLE
    r+=1

# widths
for col,w in {"A":18,"B":13,"C":9,"D":9,"E":9,"F":9,"G":9,"H":9,"I":34}.items():
    ws.column_dimensions[col].width=w
ws.freeze_panes="A4"
# ---------------- Instructions tab ----------------
ins = wb.create_sheet("Instructions", 0)   # first tab
ins.sheet_view.showGridLines = False
ins.column_dimensions["A"].width = 3
ins.column_dimensions["B"].width = 24
ins.column_dimensions["C"].width = 105
def _t(r, c, txt, **kw):
    cell = ins.cell(r, c, txt); cell.font = Font(**kw)
    cell.alignment = Alignment(vertical="top", wrap_text=True); return cell
_t(1, 2, "TolStack — How to Use", bold=True, size=18)
r_i = 3
blocks = [
 ("h", "Workflow", ""),
 ("p", "1.", "Fill in the Example tab (or copy it). Green = nominal vectors, red = random errors, blue = compensator, gold = settings, purple = results."),
 ("p", "2.", "Highlight the whole input block (for the Example tab, A4:I31)."),
 ("p", "3.", "Click Solve. Results fill the purple rows; plots embed to the right (path plot, then error histogram to its right, then sensitivity tornado)."),
 ("p", "4.", "Click Check Path to preview only the nominal coordinate path."),
 ("h", "Buttons", ""),
 ("p", "Solve", "Runs the Monte-Carlo, writes RESULT / MU / SIGMA / WORST_CASE, and embeds the plots."),
 ("p", "Check Path", "Embeds just the nominal path plot to confirm the geometry."),
 ("p", "Export Data", "In the fallback mini-window (desktop shortcut): saves the raw histogram samples to .csv or .xlsx."),
 ("h", "Input tags  (values go in the cells to the RIGHT of each tag)", ""),
 ("t", "R", "Nominal vector  X Y Z Tx Ty Tz  — one row per segment. Positions in mm, angles in radians."),
 ("t", "Re", "Random error for each R row: the +/- value at N_SIGMA (e.g. a 3-sigma tolerance)."),
 ("t", "C", "Compensator setpoint. A nonzero value marks that DOF as actively corrected."),
 ("t", "Cv", "Compensator lever vector: where the corrector sits relative to the error (creates coupled/Abbe error)."),
 ("t", "Ce", "Compensator error: the corrector's own zero-mean repeatability (+/- at N_SIGMA). 0 = perfect corrector."),
 ("t", "N_SAMPLES", "Number of Monte-Carlo trials (>= 2). More trials = smoother statistics."),
 ("t", "N_SIGMA", "How many sigma your tolerances are quoted at (usually 3)."),
 ("t", "DISTRIBUTION", "N = Normal, U = Uniform, T = Triangular (how each error is sampled)."),
 ("t", "SEED", "Blank = a new random draw every run. A number = reproducible identical results."),
 ("t", "PLOT", "1 = embed plot images into the sheet, 0 = none."),
 ("t", "SHOW", "1 = also open interactive plot windows you can rotate/zoom (not captured by the PNG)."),
 ("t", "NAME", "A label used in the plot titles."),
 ("h", "Output tags  (written by Solve; 6 values to the right of each)", ""),
 ("t", "RESULT", "mu + N*sigma  — the upper spec estimate for each DOF."),
 ("t", "MU", "Mean (systematic bias) of the propagated error."),
 ("t", "SIGMA", "1-sigma random spread."),
 ("t", "WORST_CASE", "|mu| + N*sigma  — a conservative worst case."),
 ("h", "Tips", ""),
 ("p", "•", "Angles are radians; a 0.001 rad error at a 10 mm lever adds ~0.01 mm of coupled translation."),
 ("p", "•", "If you see 'Protected View', click Enable Editing first — the backend cannot read a protected workbook."),
 ("p", "•", "Re-running Solve replaces the embedded plots (they do not pile up)."),
]
for kind, a, c in blocks:
    if kind == "h":
        r_i += 1; _t(r_i, 2, a, bold=True, size=13, color="305496"); r_i += 1
    elif kind == "t":
        _t(r_i, 2, a, bold=True, color="1F4E79"); _t(r_i, 3, c); r_i += 1
    else:
        _t(r_i, 2, a, bold=True); _t(r_i, 3, c); r_i += 1

wb.save(os.path.join(os.path.dirname(os.path.abspath(__file__)), "TolStack_Template.xlsx"))
print("template written")
