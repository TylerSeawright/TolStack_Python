"""file_io.py — openpyxl backend (cross-platform, no Excel needed).

Used for headless/batch/CI runs and for reading the shipped test workbook.
The live-Excel (COM) path lives in excel_io.py and is used by the on-sheet
buttons so it can edit the already-open workbook without file locking.
"""
from __future__ import annotations
from typing import Dict, List, Optional, Tuple
import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as XLImage


def _to_str(v) -> str:
    if v is None: return "NaN"
    if isinstance(v, bool): return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    return str(v)


def _kw(path):
    return {"keep_vba": True} if str(path).lower().endswith(".xlsm") else {}


def read_workbook(path, sheet=None, cell_range=None):
    """Return (grid_of_strings, startcell(row,col))."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    if cell_range:
        cells = ws[cell_range]
        grid = [[_to_str(c.value) for c in row] for row in cells]
        tl = ws[cell_range.split(":")[0]]
        start = (tl.row, tl.column)
    else:
        grid = [[_to_str(v) for v in row] for row in ws.iter_rows(values_only=True)]
        start = (1, 1)
    return grid, start


def write_results(path, sheet, result_cells: Dict[str, Tuple[int, int]],
                  vectors: Dict[str, object]):
    """Write each result vector to the 6 cells right of its tag cell."""
    wb = openpyxl.load_workbook(path, **_kw(path))
    ws = wb[sheet] if sheet else wb.active
    for tag, (row, col) in result_cells.items():
        vec = vectors.get(tag)
        if vec is None: continue
        for j, val in enumerate(vec, start=1):
            ws.cell(row=row, column=col + j, value=float(val))
    wb.save(path)


def embed_images(path, sheet, image_paths: List[str], anchor="K3",
                 clear_prev=True, col_step=13):
    """Insert PNGs left-to-right (path plot first, output histogram to its
    right), starting at `anchor`.

    If clear_prev, every image already on the sheet is removed first, so a
    re-run replaces the old plots instead of piling new ones on top.
    """
    wb = openpyxl.load_workbook(path, **_kw(path))
    ws = wb[sheet] if sheet else wb.active
    if clear_prev:
        ws._images = []  # openpyxl: drop previously embedded images
    col0 = column_index_from_string("".join(ch for ch in anchor if ch.isalpha()))
    row0 = int("".join(ch for ch in anchor if ch.isdigit()))
    for i, p in enumerate(image_paths):
        if not os.path.exists(p): continue
        ws.add_image(XLImage(p), f"{get_column_letter(col0 + i*col_step)}{row0}")
    wb.save(path)
